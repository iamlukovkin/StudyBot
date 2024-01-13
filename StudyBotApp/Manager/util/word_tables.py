
import subprocess
import os
import json
import warnings
import numpy as np

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from docx import Document
import pandas as pd

from config.settings import DIRECTORIES
from config.settings import FILE_PATHS


def extract_tables_from_word(doc_path):
    doc = Document(doc_path)
    tables = []
    for table in doc.tables:
        data = []
        for row in table.rows:
            row_data = []
            for cell in row.cells:
                row_data.append(cell.text)
            data.append(row_data)
        tables.append(data)
    return tables


def save_table_to_excel(table, excel_path):
    df = pd.DataFrame(table)
    df.to_excel(excel_path, index=False)


def read_excel(path: str):
    src_workbook = load_workbook(path)
    my_workbook = Workbook() 
    src_sheet = src_workbook.active
    my_sheet = my_workbook.active 

    for row in src_sheet.rows: 
        for cell in row: 
            value = cell.value
            my_coordinate = cell.coordinate 
            my_sheet[my_coordinate].value = value
    os.remove(path)
    return my_sheet


def convert_docx_to_xlsx(index, path, doc_path, excel_dir):
    file_path = doc_path + path
    table_paths = []
    tables = extract_tables_from_word(file_path)
    for i, table in enumerate(tables):
        excel_path = f'{excel_dir}file{index}.xlsx'
        index += 1
        save_table_to_excel(table, excel_path)
        table_paths.append(excel_path)
    return table_paths, index


def convert_to_list_of_lists(my_sheet):
    data = []
    for row in my_sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


def get_header_and_dates(matrix: list):
    header = matrix[1]
    groups = [cell for cell in header if cell is not None]
    temp = matrix[2:]
    dates = []
    for row in temp:
        date = row[0]
        if date is not None:
            dates.append(date)
    return header, groups, dates, temp


def find_exams(header: list, groups: list, dates: list, exams: list):
    finded_exams = []
    groups = [el for el in groups if 'группа' not in el and 'Группа' not in el]
    for group in groups:
        group_index = header.index(group)
        group = 'gr_' + group
        for row in exams:
            exam = row[group_index]
            date = row[0]
            if '.'  in date:
                month = date.split('.')[1]
            else:
                date = date.strip() + '.' + month.strip()
            if exam is not None:
                time = exam.split(' ', 1)[0]
                exam = exam.replace(time, '')
                if len(time) == 3:
                    time = '0' + time
                formatted_time = time[:2] + ":" + time[2:]
                finded_exams.append([group, date, formatted_time, exam])
    return finded_exams


def parse_exams(doc_path, excel_dir, save_path):
    documents = os.listdir(doc_path)
    table_paths = []
    index = 0
    exams_list = []
    for docx_path in documents:
        if not docx_path.endswith('.docx'):
            continue
        table_paths, index = convert_docx_to_xlsx(index, docx_path, doc_path, excel_dir)
        for path in table_paths:
            my_sheet = read_excel(path)
            matrix = convert_to_list_of_lists(my_sheet)
            header, groups, dates, exams = get_header_and_dates(matrix)
            exams_list.extend(find_exams(header, groups, dates, exams))
    with open(save_path, 'w') as f:
        json.dump({'Exams':exams_list}, f, ensure_ascii=False)
    return exams_list


if __name__ == '__main__':
    doc_path = DIRECTORIES['WORD_LESSONS_DIR']
    excel_dir = DIRECTORIES['EXCEL_LESSONS_DIR']
    save_path = FILE_PATHS['EXAMS_JSON']
    parse_exams(doc_path, excel_dir)
