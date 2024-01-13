import os
import warnings
import numpy as np

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def refresh_graphs(directory: str):
    """
    Refreshes graphs.
    
    Args:
        directory (str): directory path
        dst (str): destination path
        
    Returns:
        list: list of lessons
    """
    lessons = []
    warnings.filterwarnings("ignore", category=UserWarning)
    for filename in os.listdir(directory):
        my_sheet = read_excel(directory + filename)
        matrix = convert_to_list_of_lists(my_sheet)
        cropped_matrix = crop_matrix(matrix)
        finded_lessons = find_lesson(cropped_matrix)
        [lessons.append(lesson) for lesson in finded_lessons if None not in lesson]
    return lessons


def read_excel(path: str):
    src_workbook = load_workbook(path)
    my_workbook = Workbook() 
    src_sheet = src_workbook.active
    my_sheet = my_workbook.active 
    merged_cells = src_sheet.merged_cells

    for row in src_sheet.rows: 
        for cell in row: 
            value = cell.value
            my_coordinate = cell.coordinate
            if my_coordinate not in merged_cells: 
                my_sheet[my_coordinate].value = value
            else: 
                for merged_range in merged_cells:
                    if my_coordinate in merged_range:
                        for my_cell in merged_range.cells:
                            coordinate = get_column_letter(my_cell[1]) + str(my_cell[0]) 
                            if value is not None: 
                                my_sheet[coordinate].value = value
    os.remove(path) 
    return my_sheet


def convert_to_list_of_lists(my_sheet):
    data = []
    for row in my_sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


def crop_matrix(matrix: list):
    index_start = matrix.index([row for row in matrix if row[0] == 'Дни'][0])
    matrix = matrix[index_start:]
    for row in matrix:
        for cell in row:
            if type(cell) == type(np.nan):
                cell = None
    return matrix


def find_lesson(matrix: list):
    header = matrix[0]
    finded_lessons = []
    for group in header[3:]:
        if group is None:
            continue
        temp = group
        temp = temp.replace(' ', '')
        temp = temp.replace('.', '')
        temp = temp.split('смена')[0]
        index = header.index(group)
        for row in matrix:
            lesson = row[index]
            if lesson is None or lesson == '':
                continue
            day, time, week = get_lesson_info(row, matrix)
            if None not in (day, time, week):
                finded_lessons.append([temp, day, time, week, lesson])
    return finded_lessons


def get_lesson_info(row: list, matrix: list):   
    day, time, week = row[0], row[1], row[2]
    if None in (day, time, week):
        return None, None, None
    day = day.replace(' ', '')
    time = time.replace(' ', '')
    week = week.replace(' ', '')
    week = 'Знаменатель' if week == 'Знам.' else 'Числитель'
    return day.capitalize(), time, week
